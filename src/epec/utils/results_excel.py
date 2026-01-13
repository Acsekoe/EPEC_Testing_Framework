from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:  # optional dependency
    Workbook = None
    Font = None
    Alignment = None
    get_column_letter = None


def _safe_float(x: Any) -> Any:
    """Convert numeric-ish objects to float when possible (Pyomo values, numpy scalars, etc.)."""
    try:
        # pyomo value-like
        if hasattr(x, "value"):
            x = x.value
        # numpy scalar-like
        if hasattr(x, "item") and callable(x.item):
            x = x.item()
        # plain numeric
        if isinstance(x, (int, float)):
            return float(x)
        return x
    except Exception:
        return x


def _auto_fit_columns(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
        values = ["" if v is None else str(v) for v in col]
        width = min(max((len(v) for v in values), default=0) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)


def _write_kv_sheet(ws, title: str, data: Dict[str, Any]) -> None:
    ws.title = title
    ws.append(["key", "value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for k in sorted(data.keys()):
        ws.append([k, _safe_float(data[k])])
    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _theta_to_dict(theta: Any) -> Dict[str, Any]:
    """
    Best-effort conversion. Works if theta has .q_man/.d_offer/.T dict-like.
    """
    out: Dict[str, Any] = {}
    for name in ("q_man", "d_offer", "T"):
        if hasattr(theta, name):
            out[name] = getattr(theta, name)
    return out


def _flatten_region_dict(prefix: str, regions: Iterable[str], d: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    d = d or {}
    return {f"{prefix}_{r}": _safe_float(d.get(r, None)) for r in regions}


def _flatten_arc_dict(prefix: str, arcs: Iterable[Tuple[str, str]], d: Optional[Dict[Tuple[str, str], Any]]) -> Dict[str, Any]:
    d = d or {}
    out: Dict[str, Any] = {}
    for (e, r) in arcs:
        out[f"{prefix}_{e}_{r}"] = _safe_float(d.get((e, r), None))
    return out


def save_run_results_excel(
    project_root: Path,
    *,
    sets: Any,
    params: Any,
    theta_star: Any,
    hist: List[Dict[str, Any]],
    run_cfg: Dict[str, Any],
    ipopt_opts: Optional[Dict[str, Any]] = None,
    filename_prefix: str = "run_small",
) -> Path:
    """
    Writes an Excel file to <project_root>/results/<prefix>_<timestamp>_....xlsx

    What it logs into 'history' (best-effort, based on what's present in hist rows):
      - iter, region, accepted, status, term, ulp_obj, llp_obj
      - lambda_* (module-balance duals)
      - pi_*     (production-balance duals)
      - u_short_* and nu_ushort_* (optional feasibility slack + its dual)
      - x_man_*
      - x_flow_e_r for all pairs (including diagonal)
      - br_T_in_e_r for trade arcs (tariff best-response into the current region)

    If keys are missing in hist rows, corresponding columns will be empty.
    """
    if Workbook is None:
        raise ModuleNotFoundError(
            "openpyxl is required to write Excel results. Install it with: pip install openpyxl"
        )

    project_root = Path(project_root)
    results_dir = project_root / "results"
    results_dir.mkdir(parents=True, exist_ok=True)

    # nice filename with key params
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    eps = run_cfg.get("eps", None)
    eps_u = run_cfg.get("eps_u", None)
    tol = run_cfg.get("tol", None)
    max_iter = run_cfg.get("max_iter", None)
    damping = run_cfg.get("damping", None)
    price_sign = run_cfg.get("price_sign", None)

    def _fmt(v: Any) -> str:
        if v is None:
            return "NA"
        try:
            return f"{float(v):.0e}"
        except Exception:
            return str(v)

    fname = (
        f"{filename_prefix}_{ts}"
        f"_eps={_fmt(eps)}_eps_u={_fmt(eps_u)}_tol={_fmt(tol)}"
        f"_max_iter={_fmt(max_iter)}_damping={_fmt(damping)}_price_sign={_fmt(price_sign)}.xlsx"
    )
    out_path = results_dir / fname

    R = list(getattr(sets, "R"))
    RR = list(getattr(sets, "RR"))
    RRx = list(getattr(sets, "RRx", RR))

    wb = Workbook()

    # --- run_config sheet
    ws_cfg = wb.active
    cfg_data = {
        **{f"run_cfg.{k}": v for k, v in (run_cfg or {}).items()},
        **{f"ipopt.{k}": v for k, v in (ipopt_opts or {}).items()},
    }
    _write_kv_sheet(ws_cfg, "run_config", cfg_data)

    # --- final_theta sheet
    ws_theta = wb.create_sheet("final_theta")
    ws_theta.append(["category", "key", "value"])
    for c in ("A1", "B1", "C1"):
        ws_theta[c].font = Font(bold=True)

    theta_dict = _theta_to_dict(theta_star)
    # q_man, d_offer
    for k in ("q_man", "d_offer"):
        dd = theta_dict.get(k, {}) or {}
        for r in R:
            ws_theta.append([k, r, _safe_float(dd.get(r, None))])
    # T arcs
    T = theta_dict.get("T", {}) or {}
    for (e, r) in RR:
        ws_theta.append(["T", f"{e}->{r}", _safe_float(T.get((e, r), None))])

    ws_theta.freeze_panes = "A2"
    _auto_fit_columns(ws_theta)


    # --- history sheet
    ws_h = wb.create_sheet("history")

    # build columns
    base_cols = ["iter", "region", "accepted", "status", "term", "ulp_obj", "llp_obj"]

    # Region-level diagnostics / quantities
    region_cols: List[str] = []
    for pfx in ["lambda", "pi", "u_short", "nu_ushort", "x_man"]:
        region_cols.extend([f"{pfx}_{r}" for r in R])

    # ULP strategic variables snapshot (every iteration)
    region_cols.extend([f"q_man_{r}" for r in R])
    region_cols.extend([f"d_offer_{r}" for r in R])

    # Arc-level quantities
    arc_cols: List[str] = []

    # All flows (including diagonal)
    arc_cols.extend([f"x_flow_{e}_{r}" for (e, r) in RRx])

    # Strategic trade policy snapshots (Option B)
    arc_cols.extend([f"tau_{e}_{r}" for (e, r) in RR])
    arc_cols.extend([f"markup_{e}_{r}" for (e, r) in RR])

    # Best-response decisions for the active player (optional)
    arc_cols.extend([f"br_tau_in_{e}_{r}" for (e, r) in RR])
    arc_cols.extend([f"br_m_out_{e}_{r}" for (e, r) in RR])

    # Legacy (optional): keep old columns if present in hist rows
    arc_cols.extend([f"br_T_in_{e}_{r}" for (e, r) in RR])

    extra_cols = ["max_u_short", "br_q_man", "br_d_offer"]
    cols = base_cols + region_cols + extra_cols + arc_cols

    ws_h.append(cols)
    for j in range(1, len(cols) + 1):
        cell = ws_h.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws_h.freeze_panes = "A2"
    ws_h.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    def _theta_snapshot_from_row(row: Dict[str, Any]) -> Dict[str, Any]:
        """Best-effort extraction of theta snapshots from a history row."""
        # Preferred: explicitly stored dict snapshots
        snap = {
            "q_man": row.get("theta_q_man") or row.get("q_man"),
            "d_offer": row.get("theta_d_offer") or row.get("d_offer"),
            "tau": row.get("theta_tau") or row.get("tau"),
            "markup": row.get("theta_markup") or row.get("markup") or row.get("m"),
        }
        # If nothing is present, try a nested Theta object
        if (snap["q_man"] is None and snap["d_offer"] is None and snap["tau"] is None and snap["markup"] is None):
            th = row.get("theta") or row.get("theta_star") or row.get("theta_curr")
            if th is not None:
                d = _theta_to_dict(th)
                snap["q_man"] = d.get("q_man")
                snap["d_offer"] = d.get("d_offer")
                snap["tau"] = d.get("tau")
                snap["markup"] = d.get("markup") or d.get("m")
        return snap

    # fill rows
    for row in hist:
        flat: Dict[str, Any] = {}
        flat["iter"] = row.get("iter")
        flat["region"] = row.get("region")
        flat["accepted"] = row.get("accepted")
        flat["status"] = row.get("status")
        flat["term"] = row.get("term")
        flat["ulp_obj"] = _safe_float(row.get("ulp_obj"))
        flat["llp_obj"] = _safe_float(row.get("llp_obj"))
        flat["max_u_short"] = _safe_float(row.get("max_u_short"))

        # region-level
        flat.update(_flatten_region_dict("lambda", R, row.get("lambda")))
        flat.update(_flatten_region_dict("pi", R, row.get("pi")))
        flat.update(_flatten_region_dict("u_short", R, row.get("u_short")))
        flat.update(_flatten_region_dict("nu_ushort", R, row.get("nu_ushort")))
        flat.update(_flatten_region_dict("x_man", R, row.get("x_man")))

        # theta snapshots (q_man, d_offer, tau, markup)
        snap = _theta_snapshot_from_row(row)
        flat.update(_flatten_region_dict("q_man", R, snap.get("q_man")))
        flat.update(_flatten_region_dict("d_offer", R, snap.get("d_offer")))
        flat.update(_flatten_arc_dict("tau", RR, snap.get("tau")))
        flat.update(_flatten_arc_dict("markup", RR, snap.get("markup")))

        # flows
        flat.update(_flatten_arc_dict("x_flow", RRx, row.get("x_flow")))

        # best-response (optional)
        flat.update(_flatten_arc_dict("br_tau_in", RR, row.get("br_tau_in")))
        flat.update(_flatten_arc_dict("br_m_out", RR, row.get("br_m_out")))
        flat.update(_flatten_arc_dict("br_T_in", RR, row.get("br_T_in")))
        flat["br_q_man"] = _safe_float(row.get("br_q_man"))
        flat["br_d_offer"] = _safe_float(row.get("br_d_offer"))

        ws_h.append([flat.get(c, None) for c in cols])

    _auto_fit_columns(ws_h)
    wb.save(out_path)
    return out_path